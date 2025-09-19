from datetime import datetime
import os

import pytest
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from pages.timkiem_page import Timkiem
from utils.read import read

test_data = read("Teelab.xlsx", sheet_name="Timkiem")
report_created = False

def report(filename, row_data):
    global report_created
    if not report_created:

        wb = Workbook()
        ws = wb.active
        ws.append([
            "Time",
            "Keyword",
            "Expected",
            "Actual",
            "Expected quantity",
            "Actual quantity",
            "Status"
        ])
        wb.save(filename)
        report_created = True

    wb = load_workbook(filename)
    ws = wb.active
    ws.append(row_data)
    wb.save(filename)

@pytest.mark.parametrize("tukhoa, expected", test_data)
def test_timkiem(tukhoa, expected):
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get("https://teelab.vn/")
    timkiem_page = Timkiem(driver)

    timkiem_page.tim(tukhoa)
    actual = timkiem_page.get_ketqua()
    sl_mongdoi = timkiem_page.get_slmongdoi()
    sl_thucte = timkiem_page.get_slthucte()

    print(f"Từ khóa: {tukhoa}")
    print(f"Kết quả mong đợi: {expected}")
    print(f"Kết quả thực tế: {actual}")
    print(f"Số sản phẩm mong đợi: {sl_mongdoi}")
    print(f"Số sản phẩm thực tế: {sl_thucte}")

    test_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    filename = os.path.join("tests", "report", "Timkiem_Report.xlsx")
    assert expected.strip() == actual.strip(), f"Kết quả mong đợi {sl_mongdoi}, thực tế {sl_thucte}"
    assert sl_thucte == sl_mongdoi, f"Số sản phẩm mong đợi {sl_mongdoi}, thực tế {sl_thucte}"

    status = "PASS"
    report(filename,[test_time, tukhoa, expected, actual, sl_mongdoi, sl_thucte, status])
    driver.quit()