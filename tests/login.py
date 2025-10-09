import os
from datetime import datetime

import pytest
from openpyxl import Workbook, load_workbook
from selenium import webdriver

from pages.login_page import Login
from utils.data_untils import read

test_data = read("Teelab.xlsx", sheetname="Login")
report_created = False
ids = [f"{i+1}. ({row[2]})" for i, row in enumerate(test_data)]

def report(filename, row_data):
    global report_created
    if not report_created:

        wb = Workbook()
        ws = wb.active
        ws.append([
            "Time",
            "Email",
            "Password",
            "Expected",
            "Actual",
            "Status"
        ])
        wb.save(filename)
        report_created = True

    wb = load_workbook(filename)
    ws = wb.active
    ws.append(row_data)
    wb.save(filename)

@pytest.mark.parametrize("email, password, expected", test_data, ids=ids)
def test_dangnhap(email, password, expected):
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get("https://teelab.vn/")
    login_page = Login(driver)

    login_page.account()
    login_page.email_enter(email)
    login_page.password_enter(password)
    login_page.login()

    actual = login_page.get_result()

    print(f"Kết quả mong đợi: {expected}")
    print(f"Kết quả thực tế: {actual}")

    test_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    filename = os.path.join("tests", "reports", "Login_Report.xlsx")

    try:
        assert actual.strip() == expected.strip(), f"Expected: {expected}, Actual: {actual}"
        status = "PASS"
    except AssertionError:
        status = "FAIL"
        raise

    finally:
        report(filename, [test_time, email, password, expected, actual, status])
        driver.quit()