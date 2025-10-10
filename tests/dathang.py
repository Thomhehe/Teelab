import os
from datetime import datetime

import pytest
from openpyxl import Workbook, load_workbook
from selenium import webdriver

from pages.dathang_page import Dathang
from utils.data_untils import load_excel_data

test_data = load_excel_data("Teelab.xlsx", sheetname="Dathang")
report_created = False

def report(filename, row_data):
    global report_created
    if not report_created:

        wb = Workbook()
        ws = wb.active
        ws.append([
            "Thời gian",
            "Họ tên",
            "Số điện thoại",
            "Địa chỉ",
            "Tình thành",
            "Quận huyện",
            "Phường xã",
            "Kết quả mong đợi",
            "Kết quả thực tế",
            "Trạng thái"
        ])
        wb.save(filename)
        report_created = True

    wb = load_workbook(filename)
    ws = wb.active
    ws.append(row_data)
    wb.save(filename)

@pytest.mark.parametrize("hoten, sdt, diachi, tinhthanh, quanhuyen, phuongxa, expected", test_data)
def test_dathang(hoten, sdt, diachi, tinhthanh, quanhuyen, phuongxa, expected):
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get("https://teelab.vn/")

    dathang_page = Dathang(driver)
    dathang_page.dathang(hoten, sdt, diachi, tinhthanh, quanhuyen, phuongxa)

    actual = dathang_page.lay_thongbao()

    print(f"Kết quả mong đợi: {expected}")
    print(f"Kết quả thực tế: {actual}")

    test_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    filename = os.path.join("tests", "reports", "Dathang_Report.xlsx")

    try:
        assert expected.strip() == actual.strip(), f"Thông báo mong đợi {expected}, thực tế {actual}"
        status = "PASS"
    except AssertionError:
        status = "FAIL"
        raise
    finally:
        report(filename, [test_time, hoten, sdt, diachi, tinhthanh, quanhuyen, phuongxa, expected, actual, status])
        driver.quit()