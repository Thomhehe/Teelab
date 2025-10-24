import os
from datetime import datetime

import pytest
from openpyxl import Workbook, load_workbook
from selenium import webdriver

from pages.login_page import Login
from utils.data_utils import load_excel_data
from utils.screenshot_utils import take_screenshot

test_data = load_excel_data("Teelab.xlsx", sheetname="Login")
ids = [f"{i+1}. ({row[2]})" for i, row in enumerate(test_data)]

filename_report = r"D:\PyCharm\Teelab\reports\Login_Report.xlsx"
if os.path.exists(filename_report):
    os.remove(filename_report)

def report(filename, row_data):
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Time",
            "Email",
            "Password",
            "Expected",
            "Actual",
            "Status",
            "Screenshot"
        ])
    else:
        wb = load_workbook(filename)
        ws = wb.active

    ws.append(row_data)
    wb.save(filename)

@pytest.mark.parametrize("email, password, expected", test_data, ids=ids)
def test_login(email, password, expected):
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get("https://teelab.vn/")
    login_page = Login(driver)

    login_page.account()
    login_page.email_enter(email)
    login_page.password_enter(password)
    login_page.login()

    actual = login_page.get_result()

    print(f"\nExpected: {expected}")
    print(f"Actual: {actual}")

    test_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    screenshot_path = ""

    try:
        assert actual.strip() == expected.strip(), f"Expected: {expected}, Actual: {actual}"
        status = "PASS"
    except AssertionError:
        status = "FAIL"
        screenshot_path = take_screenshot(driver, f"loginfail_{email}")
        raise

    finally:
        report(filename_report, [test_time, email, password, expected, actual, status, screenshot_path])
        driver.quit()