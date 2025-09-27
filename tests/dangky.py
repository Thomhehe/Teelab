import os
from datetime import datetime

import pytest
from openpyxl import Workbook, load_workbook
from selenium import webdriver

from pages.dangky_page import Dangky
from utils.read import read

test_data = read("Teelab.xlsx", sheet_name="Dangky")
report_created = False

ids = [f"{i+1}. ({row[5]})" for i, row in enumerate(test_data)]

def report(filename, row_data):
    global report_created
    if not report_created:

        wb = Workbook()
        ws = wb.active
        ws.append([
            "Time",
            "LastName",
            "Name",
            "Email",
            "Phone",
            "Password",
            "Expected",
            "Actual",
            "Status"
        ])
        wb.save(filename)
        report_created = True

    wb = load_workbook(filename)
    ws = wb.active
    ws.append(row_data)
    wb.save(filename)

@pytest.mark.parametrize("ho, ten, email, sdt, matkhau, expected", test_data, ids=ids)
def test_dangky(ho, ten, email, sdt, matkhau, expected):
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get("https://teelab.vn/")
    dangky_page = Dangky(driver)

    dangky_page.dangky(ho, ten, email, sdt, matkhau)

    actual = dangky_page.lay_thongbao()

    print(f"Kết quả mong đợi: {expected}")
    print(f"Kết quả thực tế: {actual}")

    test_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    filename = os.path.join("tests", "reports", "Dangky_Report.xlsx")

    try:
        assert actual.strip() == expected.strip(), f"Expected: {expected}, Actual: {actual}"
        status = "PASS"
    except AssertionError:
        status = "FAIL"
        raise

    finally:
        report(filename, [test_time, ho, ten, email, matkhau, expected, actual, status])
        driver.quit()