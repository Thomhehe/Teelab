import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from selenium import webdriver

from pages.giohang_page import Giohang

report_created = False

def report(filename, row_data):
    global report_created
    if not report_created:

        wb = Workbook()
        ws = wb.active
        ws.append([
            "Time",
            "Product",
            "Price",
            "Color/Size",
            "Status"
        ])
        wb.save(filename)
        report_created = True

    wb = load_workbook(filename)
    ws = wb.active
    ws.append(row_data)
    wb.save(filename)

def test_themsp():
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get("https://teelab.vn/")

    giohang_page = Giohang(driver)

    giohang_page.chon_sp_mua()
    ctsp = giohang_page.lay_thongtin_ctsp()
    ctsp_mau_kichthuoc = f"{ctsp['mausac']} / {ctsp['kichthuoc']}"
    giohang_page.them_giohang()
    gh = giohang_page.lay_thongtin_gh()

    expected = f"{ctsp['ten']} | {ctsp['gia']} | {ctsp_mau_kichthuoc}".strip()
    actual = f"{gh['ten']} | {gh['gia']} | {gh['mau_kichthuoc']}".strip()

    print(f"Thông tin mong đợi: {expected}")
    print(f"Thông tin thực tế: {actual}")

    test_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    filename = os.path.join("tests", "reports", "Giohang_Report.xlsx")

    try:
        assert expected == actual, f"Thông tin mong đợi {expected}, thực tế {actual}"
        status = "PASS"
    except AssertionError as e:
        status = "FAIL"
        raise e
    finally:
        report(filename, [test_time, ctsp['ten'], ctsp['gia'], ctsp_mau_kichthuoc, status])
        driver.quit()