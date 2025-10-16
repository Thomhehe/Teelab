import os
from datetime import datetime

import pytest
from openpyxl import Workbook, load_workbook
from selenium import webdriver

from pages.login_page import Login
from pages.signup_page import Signup
from utils.data_untils import load_excel_data
from utils.screenshot_utils import take_screenshot

test_data = load_excel_data("Teelab.xlsx", sheetname="Signup")
report_created = False

ids = [f"{i+1}. ({row[5]})" for i, row in enumerate(test_data)]

def report(filename, row_data):
    global report_created
    if not report_created:

        wb = Workbook()
        ws = wb.active
        ws.append([
            "Time",
            "LastName",
            "Name",
            "Email",
            "Phone",
            "Password",
            "Expected",
            "Actual",
            "Status",
            "Screenshot"
        ])
        wb.save(filename)
        report_created = True

    wb = load_workbook(filename)
    ws = wb.active
    ws.append(row_data)
    wb.save(filename)

@pytest.mark.parametrize("lastname, name, email, phone, password, expected", test_data, ids=ids)
def test_signup(lastname, name, email, phone, password, expected):
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get("https://teelab.vn/")
    login_page = Login(driver)
    signup_page = Signup(driver)

    login_page.account()
    signup_page.signup_select()
    signup_page.lastname_enter(lastname)
    signup_page.name_enter(name)
    signup_page.email_enter(email)
    signup_page.phone_enter(phone)
    signup_page.password_enter(password)
    signup_page.signup_enter()

    actual = signup_page.get_result()

    print(f"Expected: {expected}")
    print(f"Actual: {actual}")

    test_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    filename = os.path.join("tests", "reports", "Signup_Report.xlsx")
    screenshot_path = ""

    try:
        assert actual.strip() == expected.strip(), f"Expected: {expected}, Actual: {actual}"
        status = "PASS"
    except AssertionError:
        status = "FAIL"
        screenshot_path = take_screenshot(driver, f"signupfail_{email}")
        raise

    finally:
        report(filename, [test_time, lastname, name, email, phone, password, expected, actual, status, screenshot_path])
        driver.quit()