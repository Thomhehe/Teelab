import os
import time
from datetime import datetime
from selenium import webdriver
from openpyxl import load_workbook, Workbook

from pages.editcart_page import EditCart
from utils.data_untils import load_excel_data
import pytest

test_data = load_excel_data("Teelab.xlsx", sheetname="EditQuantity")

formatted_data = []
ids = []
for i, row in enumerate(test_data):
    try:
        action_ = str(row[0]).strip().lower()
        value_ = int(row[1])
        formatted_data.append((action_, value_))
        ids.append(f"{i+1}. ({action_} {value_})")
    except Exception as e:
        print(f"Bỏ qua dòng dữ liệu lỗi: {row} ({e})")

report_created = False

def report(filename, row_data):
    global report_created
    if not report_created:
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Time",
            "Action",
            "Value",
            "Expected_Qty",
            "Actual_Qty",
            "Expected_Total",
            "Actual_Total",
            "Status"
        ])
        wb.save(filename)
        report_created = True

    wb = load_workbook(filename)
    ws = wb.active
    ws.append(row_data)
    wb.save(filename)


@pytest.mark.parametrize("action, value", formatted_data, ids=ids)
def test_editcart(action, value):
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get("https://teelab.vn/")

    editcart_page = EditCart(driver)
    editcart_page.select_product_buy()
    editcart_page.add_product_cart()

    time.sleep(2)

    old_qty = editcart_page.get_quantity()

    if action == "decrease" and old_qty <= value:
        print("Số lượng quá thấp, tự động tăng trước để tránh xóa sản phẩm.")
        editcart_page.increase_quantity(times=value + 1)
        time.sleep(1)
        old_qty = editcart_page.get_quantity()

    if action == "increase":
        editcart_page.increase_quantity(times=value)
        expected_qty = old_qty + value
    elif action == "decrease":
        editcart_page.decrease_quantity(times=value)
        expected_qty = max(1, old_qty - value)
    elif action == "input_quantity":
        editcart_page.input_quantity(qty=value)
        expected_qty = value
    else:
        pytest.skip(f"Hành động không hợp lệ: {action}")
    time.sleep(2)

    new_qty = editcart_page.get_quantity()
    new_total = editcart_page.get_total_amount()

    try:
        actual_total, expected_total = editcart_page.verify_total_amount()
        status = "PASS" if new_qty == expected_qty and actual_total == expected_total else "FAIL"
    except AssertionError as e:
        print(f"Lỗi xác minh tổng tiền: {e}")
        actual_total = new_total
        expected_total = editcart_page.calculate_total_amount()
        status = "FAIL"

    print(f"[{action}] Expected quantity={expected_qty}, Actual quantity={new_qty}")
    print(f"[{action}] Expected total={expected_total}, Actual total={actual_total}")

    test_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    filename_report = os.path.join("tests", "reports", "EditCart_Report.xlsx")

    report(filename_report, [
        test_time,
        action,
        value,
        expected_qty,
        new_qty,
        expected_total,
        actual_total,
        status
    ])

    driver.quit()
