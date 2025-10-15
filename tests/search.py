from datetime import datetime
import os

import pytest
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from pages.search_page import Search
from utils.data_untils import load_excel_data

test_data = load_excel_data("Teelab.xlsx", sheetname="Search")
report_created = False
ids = [f"{i+1}. ({row[0]})" for i, row in enumerate(test_data)]

def report(filename, row_data):
    global report_created
    if not report_created:

        wb = Workbook()
        ws = wb.active
        ws.append([
            "Time",
            "Keyword",
            "Expected",
            "Actual",
            "Expected Quantity",
            "Actual Quantity",
            "Status"
        ])
        wb.save(filename)
        report_created = True

    wb = load_workbook(filename)
    ws = wb.active
    ws.append(row_data)
    wb.save(filename)

@pytest.mark.parametrize("keyword, expected", test_data, ids=ids)
def test_search(keyword, expected):
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get("https://teelab.vn/")
    search_page = Search(driver)

    search_page.search(keyword)
    actual = search_page.get_result()
    quantity_expected = search_page.get_quantity_expected()
    quantity_actual = search_page.get_quantity()

    print(f"Keyword: {keyword}")
    print(f"Expected: {expected}")
    print(f"Actual: {actual}")
    print(f"Quantity expected: {quantity_expected}")
    print(f"Quantity actual: {quantity_actual}")

    test_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    filename = os.path.join("tests", "reports", "Search_Report.xlsx")

    try:
        assert expected.strip() == actual.strip(), f"Expected {expected}, Actual {actual}"
        assert quantity_actual == quantity_expected, f"Quantity expected {quantity_expected}, actual {quantity_actual}"
        status = "PASS"
    except AssertionError:
        status = "FAIL"
        raise
    finally:
        report(filename, [test_time, keyword, expected, actual, quantity_expected, quantity_actual, status])
        driver.quit()