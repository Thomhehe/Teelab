import os
from datetime import datetime

import pytest
from openpyxl import Workbook, load_workbook
from selenium import webdriver

from pages.dangnhap_page import Dangnhap
from utils.read import read

test_data = read("Teelab.xlsx", sheet_name="Dangnhap")
report_created = False

def report(filename, row_data):
    global report_created
    if not report_created:

        wb = Workbook()
        ws = wb.active
        ws.append([
            "Time",
            "Email",
            "Password",
            "Expected",
            "Actual",
            "Status"
        ])
        wb.save(filename)
        report_created = True

    wb = load_workbook(filename)
    ws = wb.active
    ws.append(row_data)
    wb.save(filename)

@pytest.mark.parametrize("email, matkhau, expected", test_data)
def test_dangnhap(email, matkhau, expected):
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get("https://teelab.vn/account/login")
    dangnhap_page = Dangnhap(driver)

    dangnhap_page.dangnhap(email, matkhau)

    actual = dangnhap_page.get_thongbao()

    test_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    filename = os.path.join("tests", "report", "Dangnhap_Report.xlsx")

    try:
        assert actual.strip() in expected.strip(), f"Expected: {expected}, Actual: {actual}"
        status = "PASS"
    except AssertionError:
        status = "FAIL"
        raise

    finally:
        report(filename, [test_time, email, matkhau, expected, actual, status])
        driver.quit()