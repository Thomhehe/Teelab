import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from selenium import webdriver

from pages.addcart_page import Cart
from utils.screenshot_utils import take_screenshot

filename_report = r"D:\PyCharm\Teelab\reports\AddCart_Report.xlsx"
if os.path.exists(filename_report):
    os.remove(filename_report)

def report(filename, row_data):
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Time",
            "Product_expected",
            "Product_actual",
            "Price_expected",
            "Price_actual",
            "Color/Size_expected",
            "Color/Size_actual",
            "Total_expected",
            "Total_actual",
            "Status",
            "Screenshot"
        ])
    else:
        wb = load_workbook(filename)
        ws = wb.active

    ws.append(row_data)
    wb.save(filename)

def test_addcart():
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get("https://teelab.vn/")

    addcart_page = Cart(driver)

    addcart_page.select_product_buy()
    details = addcart_page.get_information_details()
    details_color_size = f"{details['color']} / {details['size']}"
    addcart_page.add_product_cart()
    cart = addcart_page.get_information_cart()

    expected = f"{details['name']} | {details['price']} | {details_color_size}".strip()
    actual = f"{cart['name']} | {cart['price']} | {cart['color_size']}".strip()

    print(f"\nExpected: {expected}")
    print(f"Actual: {actual}")

    # --- Tính tổng tiền mong đợi & lấy tổng tiền thực tế ---
    total_expected = addcart_page.calculate_total_amount()
    total_actual = addcart_page.get_total_amount()

    print(f"Total_expected: {total_expected}")
    print(f"Total_actual: {total_actual}")

    test_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    screenshot_path = ""

    try:
        assert expected == actual, f"Expected{expected}, actual {actual}"
        assert total_expected == total_actual, f"Expected{total_expected}, actual {total_actual}"
        status = "PASS"
    except AssertionError as e:
        status = "FAIL"
        screenshot_path = take_screenshot(driver, f"addcartfail_{details['name']}")
        raise e
    finally:
        report(filename_report, [test_time, details['name'], cart['name'], details['price'], cart['price'],details_color_size, cart['color_size'], total_expected, total_actual, status, screenshot_path])
        driver.quit()