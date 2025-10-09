import csv
import json
import openpyxl

def load_csv_data(filepath):
    with open(filepath, newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        return list(reader)

def load_json_data(filepath):
    with open(filepath, encoding='utf-8') as jsonfile:
        return json.load(jsonfile)

def load_excel_data(filepath, sheetname=None):
    workbook = openpyxl.load_workbook(filepath)

    if sheetname:
        if sheetname in workbook.sheetnames:
            sheet = workbook[sheetname]
        else:
            raise ValueError(f"Sheet '{sheetname}' not found in the Excel file.")
    else:
        if workbook.sheetnames:
            sheet = workbook[workbook.sheetnames[0]]
        else:
            raise ValueError("No sheets found in the Excel file.")

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(list(row))

    return data