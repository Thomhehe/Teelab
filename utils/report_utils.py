import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

def write_report(filename, data_dict):
    # Tạo thư mục nếu chưa có
    os.makedirs(os.path.dirname(filename), exist_ok=True)

    # Thêm thời gian hiện tại nếu chưa có
    if "Time" not in data_dict:
        data_dict["Time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Sắp xếp lại để "Time" luôn là cột đầu tiên
    ordered_data = {"Time": data_dict.pop("Time")}
    ordered_data.update(data_dict)

    # Nếu file chưa tồn tại → tạo file mới
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        ws.append(list(ordered_data.keys()))   # header
        ws.append(list(ordered_data.values())) # data
        wb.save(filename)
        wb.close()
        return

    # Nếu file đã tồn tại → load lên và xóa dữ liệu cũ
    wb = load_workbook(filename)
    ws = wb.active

    # Xóa tất cả dữ liệu cũ (bao gồm cả header)
    ws.delete_rows(1, ws.max_row)

    # Ghi lại header và dòng dữ liệu mới
    ws.append(list(ordered_data.keys()))
    ws.append(list(ordered_data.values()))

    wb.save(filename)
    wb.close()
